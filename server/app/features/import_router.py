"""
CSV / Excel collection import router.

Provides a downloadable CSV template and an upload endpoint that parses
CSV or Excel files and inserts rows into the Supabase ``items`` table.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import uuid as _uuid
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app.auth import get_current_user_id
from app.errors import error_response
from app.lib.fx_service import convert_to_eur
from app.rate_limit import per_user_rate_limit

router = APIRouter(prefix="/api/imports", tags=["Imports"])

_import_limit = per_user_rate_limit(5, window_seconds=3600, scope="collection_import")

_logger = logging.getLogger(__name__)

# Score floor for accepting a catalog match. /catalog/match documents
# >= 0.75 = strong; a WRONG canonical_key is worse than none, because it prices
# the row as a different product and looks authoritative while being incorrect.
_CANONICAL_MATCH_FLOOR = 0.75


async def _resolve_canonical_key(title, category, pool) -> "str | None":
    """
    Best-effort catalog match -> BARE canonical_key, or None.

    canonical_key is BARE (`sm10-sm10-101`), never namespaced — CLAUDE.md
    "Identifier formats". The trigger derives canonical_ref from it; never set
    canonical_ref by hand.

    Never raises: a failed match must not fail the import. The row is saved
    unpriced, which is what every imported row did before this existed.
    """
    if not title or not category or pool is None:
        return None
    try:
        from app.agents.intake.catalog_matching import _match_catalog_items

        matches = await _match_catalog_items(
            category_id=category,
            suggested_name=title,
            search_keywords=[],
            brand=None,
            set_code=None,
            pool=pool,
            extracted_attributes=None,
        )
    except Exception as e:
        _logger.warning("[import] catalog match failed for %r: %s", title, e)
        return None
    if not matches:
        return None
    best = matches[0]
    if float(best.get("match_score") or 0.0) < _CANONICAL_MATCH_FLOOR:
        return None
    return best.get("item_key") or None



# ---- Import template columns ----
# Canonical 12-column schema shared with /items-export/overview so that
# round-trips work: export → edit in Excel → re-import.
# Last column set change: 2026-04-29 (added purchase_price/_currency/_date).
IMPORT_COLUMNS = [
    "name", "category", "condition", "grade", "graded_by", "sealed",
    "purchase_price", "purchase_currency", "purchase_date",
    "estimated_value", "currency", "notes",
]

# Three example rows covering distinct categories so users see the variety
# of supported inputs (TCG-graded, sealed LEGO set, watch with purchase
# history). All optional except `name` + `category`.
IMPORT_EXAMPLE_ROWS = [
    {
        "name": "EXAMPLE – Charizard Base Set Holo 1st Edition (delete this row)",
        "category": "pokemon",
        "condition": "Near Mint",
        "grade": "PSA 9",
        "graded_by": "PSA",
        "sealed": "no",
        "purchase_price": "180.00",
        "purchase_currency": "EUR",
        "purchase_date": "2024-08-15",
        "estimated_value": "350.00",
        "currency": "EUR",
        "notes": "My favourite card",
    },
    {
        "name": "LEGO Star Wars UCS Millennium Falcon 75192",
        "category": "lego",
        "condition": "Mint",
        "grade": "",
        "graded_by": "",
        "sealed": "yes",
        "purchase_price": "850.00",
        "purchase_currency": "EUR",
        "purchase_date": "2023-12-26",
        "estimated_value": "1100.00",
        "currency": "EUR",
        "notes": "Sealed in original shipping carton",
    },
    {
        "name": "Rolex Submariner 116610LN",
        "category": "watches",
        "condition": "Excellent",
        "grade": "",
        "graded_by": "",
        "sealed": "no",
        "purchase_price": "7200.00",
        "purchase_currency": "USD",
        "purchase_date": "2022-04-03",
        "estimated_value": "9800.00",
        "currency": "EUR",
        "notes": "Box and papers, full service history",
    },
]


# ---- Endpoints ----

@router.get("/template")
async def import_template() -> StreamingResponse:
    """Return a CSV template with headers + 3 example rows across categories."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(IMPORT_COLUMNS)
    for ex in IMPORT_EXAMPLE_ROWS:
        writer.writerow([ex.get(c, "") for c in IMPORT_COLUMNS])
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sparrow_collect_overview.csv"},
    )


@router.post("/collection")
async def import_collection(
    file: UploadFile = File(...),
    user_id: str = Depends(get_current_user_id),
    _rl=Depends(_import_limit),
) -> Dict[str, Any]:
    """
    Accept a CSV or Excel file, parse it, and insert valid rows into
    the Supabase ``items`` table. Returns a summary with counts.
    """
    filename = file.filename or "upload"
    content = await file.read()

    if not content:
        raise error_response(400, "Empty file")

    if len(content) > 50 * 1024 * 1024:  # 50 MB max for CSV/Excel
        raise error_response(413, "File too large (max 50 MB)")

    # Decide how to parse based on extension
    lower_name = filename.lower()
    rows: List[Dict[str, Any]] = []

    try:
        if lower_name.endswith(".csv"):
            text = content.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        elif lower_name.endswith(".xlsx"):
            # openpyxl only (no pandas — too heavy for the t3.medium and pandas
            # was never installed, so this branch 500'd for every Excel upload
            # even though the FE import picker offers .xlsx).
            try:
                from openpyxl import load_workbook  # lightweight
            except ImportError:
                raise error_response(500, "Excel support is not installed on the server.")
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            row_iter = ws.iter_rows(values_only=True) if ws is not None else iter(())
            header = next(row_iter, None)
            if header:
                cols = [str(h).strip() if h is not None else "" for h in header]
                for r in row_iter:
                    if r is None or all(c is None for c in r):
                        continue
                    rows.append({cols[i]: (r[i] if i < len(r) else None)
                                 for i in range(len(cols)) if cols[i]})
            wb.close()
        elif lower_name.endswith(".xls"):
            raise error_response(400, "Legacy .xls is not supported — please re-save as .xlsx or CSV.")
        else:
            raise error_response(400, "Unsupported file type. Please upload a CSV or Excel file.")
    except HTTPException:
        raise
    except Exception as e:
        _logger.warning("Failed to parse uploaded file: %s", e)
        raise error_response(400, "Failed to parse file")

    # Normalize keys (lowercase, strip spaces)
    norm_rows: List[Dict[str, Any]] = []
    for row in rows:
        norm: Dict[str, Any] = {}
        for k, v in row.items():
            if k is None:
                continue
            norm_key = str(k).strip().lower().replace(" ", "_")
            norm[norm_key] = v
        norm_rows.append(norm)

    total = len(norm_rows)
    inserted = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []

    # Try to get DB pool for actual inserts
    pool = None
    try:
        from app.db import get_pool
        pool = get_pool()
    except (ImportError, RuntimeError, OSError) as e:
        _logger.debug("[import] DB pool unavailable, using in-memory fallback: %s", e)

    for idx, r in enumerate(norm_rows, start=1):
        name = r.get("name") or r.get("title")
        if not name or str(name).strip() == "":
            skipped += 1
            errors.append({"row": idx, "message": "Missing required 'name' column"})
            continue

        # Canonical 12-column schema (see IMPORT_COLUMNS at top of file).
        # The items table doesn't have `grade`, `graded_by`, or `sealed`
        # columns — `grade` maps to `condition_grade`; `graded_by` and
        # `sealed` go into `attributes_json`. The previous handler
        # INSERT'd into non-existent columns and would have failed on
        # any real call (latent bug; no users had imported yet).
        title = str(name).strip()
        category = str(r.get("category", "") or "").strip() or None
        condition = str(r.get("condition", "") or "").strip() or None
        condition_grade = str(r.get("grade", "") or "").strip() or None
        graded_by = str(r.get("graded_by", "") or "").strip() or None
        sealed_raw = str(r.get("sealed", "") or "").strip().lower()
        sealed = sealed_raw in ("yes", "true", "1", "y") if sealed_raw else None

        def _num(v: Any) -> Optional[float]:
            if v is None or str(v).strip() == "":
                return None
            try:
                return float(str(v).replace(",", ".").strip())
            except (ValueError, TypeError):
                return None

        purchase_price = _num(r.get("purchase_price"))
        purchase_currency = str(r.get("purchase_currency", "") or "").strip().upper() or None
        purchase_date_raw = str(r.get("purchase_date", "") or "").strip() or None
        # Best-effort YYYY-MM-DD parse; pass-through to Postgres which is
        # forgiving about ISO-ish formats. Invalid → NULL with a warning.
        purchase_date = None
        if purchase_date_raw:
            try:
                from datetime import date as _date
                purchase_date = _date.fromisoformat(purchase_date_raw[:10])
            except ValueError:
                _logger.debug("[import] purchase_date parse failed: %s", purchase_date_raw)

        # The items table deliberately carries BOTH halves of three pairs, and
        # different readers key on different halves — see the note at
        # add-manual.tsx:414-435 and scripts/audit_column_drift.py. This path
        # wrote only one half of each, so every imported row was:
        #   name NULL      -> nameless on the Home portfolio (the f9195fe bug,
        #                     fixed for add-manual/QuickScan but not here)
        #   ..._eur NULL   -> no cost basis in the analytics DCA series
        #   purchased_at   -> NULL, and that is the half ITEMS_SELECT reads
        #                     (itemsProvider.ts:136)
        # None of it threw: the readers all default to `?? 0` / 'Untitled'.
        purchase_price_eur = (
            await convert_to_eur(purchase_price, purchase_currency or "EUR")
            if purchase_price is not None
            else None
        )

        # ⚠️ `estimated_value` IS A EUR COLUMN, and the template invites a
        # currency beside it.
        #
        # The template's own header is
        # `... estimated_value,currency,notes`, and its third example row is a
        # Rolex bought in USD. Until 2026-08-19 `purchase_price` was converted
        # through `convert_to_eur` and `estimated_value` was stored RAW — so a
        # row saying "9800 USD" became `estimated_value = 9800`, and the whole
        # value chain reads that column as EUR (`item_value_v1` returns it as
        # `value_eur`). Proven on prod with a real import: 100 USD purchase
        # stored 86.39 EUR correctly, while a 200 USD estimate stored 200.
        #
        # The error scales with the currency: ~16% for USD, ~170x for JPY —
        # straight into the portfolio total, the analytics split and the
        # leaderboard, all of which present it as a market-comparable figure.
        #
        # `currency` is the estimate's currency; `purchase_currency` is the
        # purchase's. They are separate columns in the template and a row may
        # legitimately use both (bought in USD, valued in EUR).
        value_currency = str(r.get("currency", "") or "").strip().upper() or None
        est_val_raw = _num(r.get("estimated_value") or r.get("price") or r.get("value"))
        est_val = (
            await convert_to_eur(est_val_raw, value_currency or "EUR")
            if est_val_raw is not None
            else None
        )
        notes = r.get("notes")
        notes_str = str(notes).strip() if notes else None

        # graded_by + sealed live in items.attrs (jsonb) since the items
        # table has no dedicated columns for them.
        attrs: Dict[str, Any] = {}
        if graded_by:
            attrs["graded_by"] = graded_by
        if sealed is not None:
            attrs["sealed"] = sealed

        # Resolve a BARE canonical_key so imported rows are priceable. Without
        # it the item saves fine and shows "—" forever: canonical_key
        # --trg_items_canonical_ref--> canonical_ref --join--> price_predictions.
        # Best-effort and never fatal — a failed match imports the row unpriced,
        # which is exactly the old behaviour for every row.
        canonical_key = await _resolve_canonical_key(title, category, pool)

        if pool:
            try:
                async with pool.acquire() as conn:
                    await conn.execute(
                        """
                        INSERT INTO items (
                            id, user_id, name, title, category, condition,
                            condition_grade, attrs,
                            purchase_price, purchase_price_eur,
                            purchase_currency, purchase_date,
                            estimated_value, notes, canonical_key, source
                        )
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11,
                                $12, $13, $14, $15, 'import')
                        """,
                        str(_uuid.uuid4()),
                        user_id,
                        title,
                        title,
                        category,
                        condition,
                        condition_grade,
                        json.dumps(attrs) if attrs else None,
                        purchase_price,
                        purchase_price_eur,
                        purchase_currency,
                        # purchased_at is intentionally NOT bound here.
                        # asyncpg encodes a bare `date` into a timestamptz as
                        # midnight in the *host* timezone, so an Amsterdam box
                        # stored 2024-06-01 as 2024-05-31 22:00Z -- a day early
                        # for every UTC reader. trg_items_sync_paired_columns
                        # derives it from purchase_date pinned to UTC midnight.
                        purchase_date,
                        est_val,
                        notes_str,
                        canonical_key,
                    )
                inserted += 1
            except Exception as e:
                skipped += 1
                errors.append({"row": idx, "message": f"DB insert failed: {e}"})
        else:
            # No DB — count as inserted (parse-only mode)
            inserted += 1

    return {
        "total_rows": total,
        "inserted_count": inserted,
        "skipped_count": skipped,
        "errors": errors,
        "columns_detected": sorted(list(norm_rows[0].keys())) if norm_rows else [],
        "db_mode": "live" if pool else "parse_only",
    }
